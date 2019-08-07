"""
Extracting data from CCLE data obtained from Ghandi et al., 2019

The files that are needed as the input for the script:
  * CCLE_GCP.csv - obtained from the supplementary data under a different name, but it is the file that contains the global chromatin proteomics dataset

The files created from this script:
  * ccle_names.txt - the names of the CCLE cell lines corresponding to the H3 markers
  * h3marks.txt - the names of the H3 markers
  * h3_relval.txt - the values of the relative proteomics intensities obtained from Ghandi et al., 2019.
@author: Scott Campit
"""

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from itertools import chain
import string


def get_media(dictionary, key, lst):
    """
    This function requires a list of synonyms for a specific medium. It will construct a dictionary that maps the key to various synonyms of the medium.
    """
    for i in lst:
        try:
            dictionary[key].append(i)
        except KeyError:
            dictionary[key] = [i]
    return dictionary


def mapper(dict, series):
    """
    This function maps the values in a medium key to the dataframe elements to get the keys into the dataframe.
    """
    for k, v in dict.items():
        idx = series.isin(v)
        tmp = series[idx]
        tmp[idx] = k
        series[idx] = tmp
    return series


def rename_medium_to_common_IDs():
    """
    Extract will get the medium conditions that were used for the CCLE histone proteomics paper to grow the cells and classify them to simpler medium conditions. The result will be outputted as a table.
    """

    df = pd.read_csv(r'./../data/CCLE_GCP.csv')
    df = df.drop('BroadID', axis=1)

    media = pd.read_excel(r'./../data/summary.xlsx', sheet_name='Cell Line Annotations',
                          usecols=['CCLE_ID', 'Growth.Medium', 'Supplements'])
    media["New_medium"] = media["Growth.Medium"] + ' + ' + media["Supplements"]

    df = pd.merge(df, media, left_on='CellLineName', right_on='CCLE_ID')

    # Regex to remove
    df["New_medium"] = df["New_medium"].str.lower()
    df["New_medium"] = df["New_medium"].str.replace(' ', '')
    df["New_medium"] = df["New_medium"].str.replace('\"\"', '', regex=True)
    df["New_medium"] = df["New_medium"].str.replace('\"', '', regex=True)
    df["New_medium"] = df["New_medium"].str.replace(',', '')
    df["New_medium"] = df["New_medium"].str.replace('%', '')
    df["New_medium"] = df["New_medium"].str.replace('\+', '')
    df["New_medium"] = df["New_medium"].str.replace('\/+', '')
    df["New_medium"] = df["New_medium"].str.replace('\.+', '')
    df["New_medium"] = df["New_medium"].str.replace('\([^)]*\)', '')
    df["New_medium"] = df["New_medium"].str.replace('\(', '')
    df["New_medium"] = df["New_medium"].str.replace("'", '')
    df["New_medium"] = df["New_medium"].str.replace("-", '')
    df["New_medium"] = df["New_medium"].str.replace("^", '')
    df["New_medium"] = df["New_medium"].str.replace(";", '')
    df["New_medium"] = df["New_medium"].str.replace(":", '')

    # Transform once
    df["New_medium"] = df["New_medium"].str.replace('rpm', 'RPMI')
    df["New_medium"] = df["New_medium"].str.replace('waymouths', 'WAYMOUTH')
    df["New_medium"] = df["New_medium"].str.replace('l15', 'LFIFTEEN')
    df["New_medium"] = df["New_medium"].str.replace(
        'dmem:f12(1:1)', 'DMEM:FTWELVE(1:1)')
    df["New_medium"] = df["New_medium"].str.replace('imdm', 'IMDM')
    df["New_medium"] = df["New_medium"].str.replace('rpmi', 'RPMI')
    df["New_medium"] = df["New_medium"].str.replace('f-12katcc', 'FTWELVE')
    df["New_medium"] = df["New_medium"].str.replace('RPMIi-1640', 'RPMI')
    df["New_medium"] = df["New_medium"].str.replace(
        'mcdb105(1:1)medium199', 'MCDB105:M199(1:1)')
    df["New_medium"] = df["New_medium"].str.replace(
        'dmem:hamsf12', 'DMEM:FTWELVE(1:1)')
    df["New_medium"] = df["New_medium"].str.replace(
        'eaglesminimalessential', 'EMEM')
    df["New_medium"] = df["New_medium"].str.replace(
        'dulbeccosmodifiedeagles', 'DMEM')
    df["New_medium"] = df["New_medium"].str.replace('hamsf12', 'FTWELVE')
    df["New_medium"] = df["New_medium"].str.replace('hansf12', 'FTWELVE')
    df["New_medium"] = df["New_medium"].str.replace('mccoys5a', 'MCCOY5A')
    df["New_medium"] = df["New_medium"].str.replace(
        'williamsemedium', 'WILLIAMS')
    df["New_medium"] = df["New_medium"].str.replace('dme', 'DMEM')
    df["New_medium"] = df["New_medium"].str.replace(
        'dmem:f12', 'DMEM:FTWELVE(1:1)')
    df["New_medium"] = df["New_medium"].str.replace('rphihi', 'RPMI')
    df["New_medium"] = df["New_medium"].str.replace(
        "dulbeccosmemiscovesmdmhi", 'RPMI')
    df["New_medium"] = df["New_medium"].str.replace('alphamemhi', 'ALPHAMEM')
    df["New_medium"] = df["New_medium"].str.replace('RPMI2humamcsf', 'RPMI')
    df["New_medium"] = df["New_medium"].str.replace('glucose', 'GLC')
    df["New_medium"] = df["New_medium"].str.replace('pyruvate', 'PYR')
    df["New_medium"] = df["New_medium"].str.replace('glutathione', 'GSH')
    df["New_medium"] = df["New_medium"].str.replace('alphamem', 'ALPHAMEM')
    df["New_medium"] = df["New_medium"].str.replace('dulbeccosmem', 'DMEM')
    df["New_medium"] = df["New_medium"].str.replace('iscovesmdmhi', 'IMDM')
    df["New_medium"] = df["New_medium"].str.replace(
        'DMEMm:f12', 'DMEM:FTWELVE(1:1)')
    df["New_medium"] = df["New_medium"].str.replace('hamf10', 'FTEN')
    df["New_medium"] = df["New_medium"].str.replace('hamf12', 'FTWELVE')
    df["New_medium"] = df["New_medium"].str.replace('glutamine', 'GLN')
    df["New_medium"] = df["New_medium"].str.replace('emem', 'EMEM')
    df["New_medium"] = df["New_medium"].str.replace('mccoy5a', 'MCCOY5A')
    df["New_medium"] = df["New_medium"].str.replace('Wayouth', 'WAYMOUTH')
    df["New_medium"] = df["New_medium"].str.replace('waymouth', 'WAYMOUTH')
    df["New_medium"] = df["New_medium"].str.replace('puruvate', 'PYR')
    df["New_medium"] = df["New_medium"].str.replace('glutatone', 'GSH')
    df["New_medium"] = df["New_medium"].str.replace(
        'leibovitzsl-15medium', 'LFIFTEEN')
    df["New_medium"] = df["New_medium"].str.replace('hamsf10', 'FTEN')
    df["New_medium"] = df["New_medium"].str.replace('f12', 'FTWELVE')
    df["New_medium"] = df["New_medium"].str.replace('f-12', 'FTWELVE')
    df["New_medium"] = df["New_medium"].str.replace('acl4', 'ACL4')
    df["New_medium"] = df["New_medium"].str.replace('rpi-1640', 'RPMI')
    df["New_medium"] = df["New_medium"].str.replace(
        'mcdb1051:1media199', 'MCDB105:M199(1:1)')
    df["New_medium"] = df["New_medium"].str.replace('DMEMm', 'DMEM')
    df["New_medium"] = df["New_medium"].str.replace(
        'mcdb105:medium199(1:1)', 'ACL4')
    df["New_medium"] = df["New_medium"].str.replace(
        'acl-4', 'MCDB105:M199(1:1)')
    df["New_medium"] = df["New_medium"].str.replace('mem', 'MEM')
    df["New_medium"] = df["New_medium"].str.replace('alpha-MEM', 'ALPHAMEM')
    df["New_medium"] = df["New_medium"].str.replace(
        'DMEMF12(1:1)', 'DMEM:FTWELVE(1:1)')
    df["New_medium"] = df["New_medium"].str.replace(
        'DMEM:f:12', 'DMEM:FTWELVE(1:1)')
    df["New_medium"] = df["New_medium"].str.replace('dEMEM', 'DMEM')
    df["New_medium"] = df["New_medium"].str.replace(
        '(DMEM:F12=1:1)', 'DMEM:FTWELVE(1:1)')
    df["New_medium"] = df["New_medium"].str.replace(
        'mcdb105(1:1)199+', 'MCDB105:M199(1:1)')
    df["New_medium"] = df["New_medium"].str.replace('rpm+', 'RPMI')
    df["New_medium"] = df["New_medium"].str.replace('iscovesmdm+', 'IMDM')
    df["New_medium"] = df["New_medium"].str.replace(
        'mcdb105:medium199(1:1)+15%fbs+empty', 'MCDB105:M199(1:1)')
    df["New_medium"] = df["New_medium"].str.replace('rphi+', 'RPMI')
    df["New_medium"] = df["New_medium"].str.replace(
        "mcdb105:medium19915fbsempty'", 'MCDB105:M199(1:1)', regex=True)
    df["New_medium"] = df["New_medium"].str.replace(
        "mcdb105medium19915fbsempty", "MCDB105:M199(1:1)", regex=True)
    df["New_medium"] = df["New_medium"].str.replace('emptyempty', 'NAN')
    df["New_medium"] = df["New_medium"].str.replace('emptynempty', 'NAN')
    df["New_medium"] = df["New_medium"].str.replace('10fbs01mmneaa', 'FTEN')

    # Get rid of everything else
    df["New_medium"] = df["New_medium"].str.replace('\d+', '')
    df["New_medium"] = df["New_medium"].str.replace('[a-z]', '', regex=True)

    # Retransform to final version
    df["New_medium"] = df["New_medium"].str.replace('WAYMOUTH', 'Waymouth')
    df["New_medium"] = df["New_medium"].str.replace('LFIFTEEN', 'L15')
    df["New_medium"] = df["New_medium"].str.replace(
        'DMEMFTWELVEGLN', 'DMEM:F12 wGln')
    df["New_medium"] = df["New_medium"].str.replace('NAN', 'RPMI')
    df["New_medium"] = df["New_medium"].str.replace('DMEMFTWELVE', 'DMEM:F12')
    df["New_medium"] = df["New_medium"].str.replace(
        'DMEMFTWELVEFTEN', 'DMEM:F12')
    df["New_medium"] = df["New_medium"].str.replace('RPMIMEM', 'RPMI')
    df["New_medium"] = df["New_medium"].str.replace('FTWELVEMEM', 'F12')
    df["New_medium"] = df["New_medium"].str.replace('ALPHAMEMMEM', 'Alpha-MEM')
    df["New_medium"] = df["New_medium"].str.replace('EMEMMEM', 'EMEM')
    df["New_medium"] = df["New_medium"].str.replace('DMEMMEM', 'DMEM')
    df["New_medium"] = df["New_medium"].str.replace(
        'DMEMFTWELVEPYRGLN', 'DMEM:F12 wGln wPyr')
    df["New_medium"] = df["New_medium"].str.replace('DMEMGLN', 'DMEM wGln')
    df["New_medium"] = df["New_medium"].str.replace('FTWELVE', 'F12')
    df["New_medium"] = df["New_medium"].str.replace('MCCOYA', 'McCoy5A')
    df["New_medium"] = df["New_medium"].str.replace(
        'DMEMFTWELVEGLNPYR', 'DMEM:F12 wGln wPyr')
    df["New_medium"] = df["New_medium"].str.replace('DMEMGLC', 'DMEM wGlc')
    df["New_medium"] = df["New_medium"].str.replace('RPMIGLN', 'RPMI wGln')
    df["New_medium"] = df["New_medium"].str.replace('FTEN', 'F10')
    df["New_medium"] = df["New_medium"].str.replace('RPMIGLNMEM', 'RPMI wGln')
    df["New_medium"] = df["New_medium"].str.replace(
        "MCDB+", 'MCDB105:M199', regex=True)
    df["New_medium"] = df["New_medium"].str.replace('WILLIAMS', 'Williams')
    df["New_medium"] = df["New_medium"].str.replace('RPMIMEMGLN', 'RPMI wGln')
    df["New_medium"] = df["New_medium"].str.replace('EMEMPYR', 'EMEM wPyr')
    df["New_medium"] = df["New_medium"].str.replace(
        'WAYMOUTHGLN', 'Waymouth wGln')
    df["New_medium"] = df["New_medium"].str.replace('GLN', 'RPMI wGln')
    df["New_medium"] = df["New_medium"].str.replace(
        'DMEMFTWELVEMEMGLN', 'DMEM:F12 wGln')
    df["New_medium"] = df["New_medium"].str.replace('IMDMGLN', 'IMDM wGln')
    df["New_medium"] = df["New_medium"].str.replace('ACL', 'ACL4')
    df["New_medium"] = df["New_medium"].str.replace('RPMIFTWELVE', 'RPMI:F12')
    df["New_medium"] = df["New_medium"].str.replace(
        'FTWELVEMEMGLN', 'F12 wGln')
    df["New_medium"] = df["New_medium"].str.replace('EMEMFTEN', 'EMEM:F10')
    df["New_medium"] = df["New_medium"].str.replace('RPMIPYR', 'RPMI wPyr')
    df["New_medium"] = df["New_medium"].str.replace('RPMIEMEM', 'RPMI:EMEM')
    df["New_medium"] = df["New_medium"].str.replace('DMEMGLNMEM', 'DMEM wGln')
    df["New_medium"] = df["New_medium"].str.replace(
        'RPMIGLNPYR', 'RPMI wGln wPyr')
    df["New_medium"] = df["New_medium"].str.replace('DMEMIMDM', 'DMEM:IMDM')
    df["New_medium"] = df["New_medium"].str.replace('ALPHAMEM', 'AlphaMEM')
    df["New_medium"] = df["New_medium"].str.replace(
        'DMEMGLNPYR', 'DMEM wGln wPyr')
    df["New_medium"] = df["New_medium"].str.replace(
        'ALPHAMEMDMEM', 'AlphaMEM:DMEM')
    df["New_medium"] = df["New_medium"].str.replace('IMDMRPMIDMEM', 'RPMI')
    print(df.head(30))

    tmp = df["New_medium"].unique()
    tmp = pd.DataFrame(tmp)
    #df.to_csv('tmp1.csv', index=False)
    #tmp.to_csv('tmp2.csv', index=False)

# Had to manually edit some of the entries, but for the most part the algorithm is ~80% accurate in mapping.
#rename_medium_to_common_IDs()


def split_cellLine_and_tissues():
    df = pd.read_csv('GCP_proteomics_remapped.csv')
    df['Cell Line'] = df['CellLineName'].str.split('_').str[0]
    df['Tissue'] = df['CellLineName'].str.split('_', n=1).str[1]
    df['Tissue'] = df['Tissue'].str.replace('_', ' ')
    df['Tissue'] = df['Tissue'].str.title()
    df = df.drop(['CellLineName'], axis=1)
    df = df.set_index(['Cell Line', 'Tissue'])

    df.to_csv("GCP_proteomics_remapped.csv")
#split_cellLine_and_tissues()


def make_medium_xl_sheet():
    """
    """

    medium_conditions = pd.read_csv('GCP_proteomics_remapped.csv', usecols=[
                                    'Medium Condition', 'Tissue'])
    unique_conditions = medium_conditions['Medium Condition'].sort_values(
        ascending=True).unique()

    #wb = Workbook()
    #name = 'Medium_conditions.xlsx'
    #wb.save(filename = name)

    number_of_medium = medium_conditions['Medium Condition'].value_counts()
    number_of_tissues = medium_conditions['Tissue'].value_counts()
    book = load_workbook('./Medium_conditions.xlsx')
    writer = pd.ExcelWriter('./Medium_conditions.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    if "Summary" in book:
        pass
    else:
        number_of_medium.to_excel(
            writer, sheet_name="Summary", index=True, startcol=0)
        number_of_tissues.to_excel(
            writer, sheet_name="Summary", index=True, startcol=2)
    writer.save()

    for medium in unique_conditions:
        if medium in book:
            pass
        else:
            df = pd.DataFrame(
                columns=["Components", "MW", "g/L", "mM"])
        df.to_excel(writer, sheet_name=medium, index=False, header=True)
    writer.save()
#make_medium_xl_sheet()


def make_common_nutrient_id():
    """
    """
    book = load_workbook('./Medium_conditions.xlsx')
    writer = pd.ExcelWriter('./Medium_conditions.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    medium_components = []
    for sheet in writer.sheets:
        if "Summary" in sheet:
            pass
        else:
            df = pd.read_excel('./Medium_conditions.xlsx', sheet_name=sheet)
            components = df["Components"].tolist()
            medium_components.append(components)
    medium_components = list(chain.from_iterable(medium_components))
    medium_components = set(medium_components)
    medium_components = pd.DataFrame(medium_components)
    medium_components[0] = medium_components[0].sort_values(ascending=True)
    medium_components["Components"] = medium_components[0].str.replace(
        '[\(*\)]', '')

    medium_components["Components"] = medium_components["Components"].str.lower()
    medium_components["Components"] = medium_components["Components"].str.replace(
        '-', '', regex=True)
    medium_components["Components"] = medium_components["Components"].str.replace(
        '\u2022', ' ', regex=True)
    medium_components["Components"] = medium_components["Components"].str.replace(
        '2deoxydribose', 'twodeoxydribose', case=True)
    medium_components["Components"] = medium_components["Components"].str.replace(
        "adenosine 5'phosphate", 'Adenosine FivePrimePhosphate', case=True)
    medium_components["Components"] = medium_components["Components"].str.replace(
        "riboflavin 5'phosphate na", 'Riboflavin FivePrimePhosphate', case=True)
    medium_components["Components"] = medium_components["Components"].str.replace(
        "adenosine 5'triphosphate", 'Adenosine FivePrimeTriPhosphate', case=True)
    medium_components["Components"] = medium_components["Components"].str.replace(
        "menadione vitamin k3", 'Vitamin KThree', case=True)
    medium_components["Components"] = medium_components["Components"].str.replace(
        "vitamin d2 calciferol", 'Vitamin DTwo Calciferol', case=True)
    medium_components["Components"] = medium_components["Components"].str.replace(
        "vitamin b12 calciferol", 'Vitamin BTwelve', case=True)

    medium_components["Components"] = medium_components["Components"].str.replace(
        "\w*\d\w*", '')
    medium_components["Components"] = medium_components["Components"].str.replace(
        '"', '')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "hcl", '')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "dibasic", '')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "anhydrous", '')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "monobasic", '')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "hydrochloride", '')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "disodium", '')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "dihydrate", '')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "nacl", '')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "anhyd.", '')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "hemicalcium", '')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "acid", '')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "phos\.", 'phosphate')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "salt", '')

    medium_components["Components"] = medium_components["Components"].str.replace(
        "\w*\d\w*", '')
    medium_components["Components"] = medium_components["Components"].str.replace(
        '"', '')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "hcl", '')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "dibasic", '')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "anhydrous", '')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "monobasic", '')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "hydrochloride", '')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "disodium", '')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "dihydrate", '')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "nacl", '')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "anhyd.", '')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "hemicalcium", '')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "acid", '')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "phos\.", 'phosphate')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "\.", '')
    medium_components["Components"] = medium_components["Components"].str.replace(
        " kcl", '')
    medium_components["Components"] = medium_components["Components"].str.replace(
        " na", '')

    medium_components["Components"] = medium_components["Components"].str.replace(
        "freebase", '')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "salt", '')
    medium_components["Components"] = medium_components["Components"].str.replace(
        " ", '')

    medium_components["Components"] = medium_components["Components"].str.replace(
        "", '')

    # Manual mapping:
    medium_components["Components"] = medium_components["Components"].str.replace(
        "putrescine", 'Putrescine')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "sodium", 'Sodium')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "phosphate", 'Phosphate')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "RiboflavinFivePrimePhosphate", 'Alpha-D-Ribose 5-phosphate')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "calcium", 'Calcium')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "chloride", 'Chloride')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "pyridoxal", 'Pyridoxal')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "dglucosedextrose", 'D-Glucose')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "ThiaminmonoPhosphate", 'Thiamin monophosphate')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "lasparagine", 'L-Asparagine')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "iinositol", 'Myo-Inositol')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "manganese", 'Manganese')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "ribose", 'D-Ribose')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "lisoleucine", 'L-Isoleucine')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "dCalciumpantothenate", '(R)-Pantothenate')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "niacinamide", 'Nicotinamide')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "linoleic", 'Linoleic acid (all cis C18:2) n-6')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "vitaminaacetate", 'L-Asparagine')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "acetate", 'Acetate')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "magnesium", 'Magnesium')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "sulfate", 'Sulfate')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "lcysteine", 'L-Cysteine')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "lproline", 'L-Proline')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "dpantothenic", '(R)-Pantothenate')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "potassium", 'Potassium')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "twodeoxydD-Ribose", 'Deoxyribose C5H10O4')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "laspartic", 'L-aspartate')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "VitaminDTwoCalciferol", 'Vitamin D2; ergocalciferol')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "lcystine", 'L Cystine C6H12N2O4S2')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "uracil", 'Uracil')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "ammonium", 'Ammonium')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "ergocalciferol", 'Vitamin D2; ergocalciferol')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "lipoic", 'Lipoate')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "riboflavin", 'Riboflavin C17H20N4O6')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "thiamine", 'Thiamin')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "alphatocopherol", 'Alpha-Tocopherol')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "nitrate", 'Nitrate')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "bicarbonate", 'Bicarbonate')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "paraaminobenzoic", '4-Aminobenzoate')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "lserine", 'L-Serine')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "glucose", 'D-Glucose')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "follinic", 'Folate')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "llysine", 'L-Lysine')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "folic", 'Folate')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "hypoxanthine", 'Hypoxanthine')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "zinc", 'Zinc')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "adenine", 'Adenine')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "AdenosineFivePrimeTriPhosphate", 'ATP C10H12N5O13P3')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "lalanine", 'L-Alanine')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "guanosine", 'Guanosine')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "glutathionereduced", 'Reduced glutathione')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "AdenosineFivePrimePhosphate", 'AMP C10H12N5O7P')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "lthreonine", 'L-Threonine')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "pyruvate", 'Pyruvate')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "lleucine", 'L-Leucine')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "thymidine", 'Thymidine')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "cholesterol", 'Chsterol c')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "choline", 'Choline C5H14NO')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "lphenyL-Alanine", 'L-Phenylalanine')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "guanine", 'Guanine')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "lhydroxyproline", 'Trans 4 Hydroxy L proline C5H9NO3')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "lmethionine", 'L-Methionine')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "thymine", 'Thymine C5H6N2O2')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "guanine", 'Guanine')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "ribonucleosides", 'Nicotinate D-ribonucleoside')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "myoinositol", 'Myo-Inositol')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "lalanyllglutamine", 'L-alanine-L-glutamate')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "adenosine", 'Adenosine')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "xanthinena", 'Xanthine')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "lhistidine", 'L-Histidine')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "ltryptophan", 'L-Tryptophan')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "glycine", 'Glycine')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "uridine", 'Uridine')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "pyridoxine", 'Pyridoxine')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "lglutamine", 'L-Glutamine')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "lvaline", 'L-Valine')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "larginine", 'L-Arginine')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "lglutamic", 'L-Glutamate')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "cytidine", 'Cytidine')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "ascorbic", 'L-Ascorbate')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "biotin", 'Biotin')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "nicotinicniacin", 'Nicotinate')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "d\+galactose", 'D-Galactose')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "molybdate", 'Molybdate')
    medium_components["Components"] = medium_components["Components"].str.replace(
        "ltyrosine", 'L-Tyrosine')
    medium_components["Components"] = medium_components["Components"].str.replace(
            "vitamin", 'Cob(I)alamin')
    medium_components["Components"] = medium_components["Components"].str.replace(
            "VitaminKThree", 'Vitamin K3')

    medium_components["Components"] = medium_components["Components"].str.replace(
        "\r", '')

    medium_components["Components"] = medium_components["Components"].str.replace(
        r'(\w)([A-Z][a-z])', r"\1 \2", regex=True)

    tmp = medium_components["Components"].str.split(
        '[A-Z][^A-Z]*', expand=True).apply(pd.Series, 1).stack()

#make_common_nutrient_id()


def get_unique_ids():
    """
    """
    df = pd.read_csv('./medium_component_map.csv')
    df = df.drop_duplicates(keep='first')
    df.to_csv('./medium_component_map.csv', index=False)
#get_unique_ids()


def map_to_recon1_xchange():
    """
    """
    pd.options.mode.chained_assignment = None
    df = pd.read_csv('./medium_component_map.csv')
    metabolites = pd.read_excel('./recon1_id.xlsx', sheet_name='Metabolites')
    reactions = pd.read_excel('./recon1_id.xlsx', sheet_name='Reactions')

    exchange_reactions = reactions[reactions['Reaction ID'].str.startswith(
        'EX_')]
    exchange_reactions['Metabolite'] = exchange_reactions['Reaction ID'].str.replace(
        'EX_', '')

    recon1_map = pd.merge(exchange_reactions, metabolites,
                          how='inner', left_on='Metabolite', right_on='Metabolite ID')
    recon1_map = recon1_map.copy(deep=True)
    full_map = pd.merge(recon1_map, df, how='inner',
                        left_on='Metabolite Name', right_on='BiGG Common Names')
    full_map = full_map.drop_duplicates(keep='first')
    full_map.to_csv('./medium_component_mapped_to_recon1.csv', index=False)
    #recon1_map.to_csv('Recon1_exchange_map.csv', index=False)


#map_to_recon1_xchange()

def drop_all_duplicates():
    """
    """
    df = pd.read_excel('./Medium_conditions.xlsx')
    df = df.drop_duplicates(keep='first')
    df.to_excel('./Medium_conditions.xlsx', index=False)


#drop_all_duplicates()


def make_medium_conditions():
    """
    """

    rpmi = pd.read_excel(r'./Medium_conditions.xlsx', sheet_name='RPMI')
    dmem = pd.read_excel(r'./Medium_conditions.xlsx', sheet_name='DMEM')
    imdm = pd.read_excel(r'./Medium_conditions.xlsx', sheet_name='IMDM')
    emem = pd.read_excel(r'./Medium_conditions.xlsx', sheet_name='EMEM')
    mcdb105 = pd.read_excel(r'./Medium_conditions.xlsx', sheet_name='MCDB105')
    m199 = pd.read_excel(r'./Medium_conditions.xlsx', sheet_name='M199')
    f12 = pd.read_excel(r'./Medium_conditions.xlsx', sheet_name='F12')

    book = load_workbook(r'./Medium_conditions.xlsx')
    writer = pd.ExcelWriter(r'./Medium_conditions.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    def make_medium(medium1, medium2):
        medium1 = medium1.copy(deep=True)
        medium1['g/L'] = medium1['g/L'].replace('Infinity', np.inf)
        medium1['g/L'] = medium1['g/L']*0.5

        medium2 = medium2.copy(deep=True)
        medium2['g/L'] = medium2['g/L'].replace('Infinity', np.inf)
        medium2['g/L'] = medium2['g/L']*0.5

        combined_medium = pd.concat([medium1, medium2]).groupby(
            'Components')['g/L'].sum().reset_index()

        return combined_medium

    # DMEM-IMDM
    dmem_imdm = make_medium(dmem, imdm)
    dmem_imdm.to_excel(writer, sheet_name='DMEM-IMDM', index=False)

    # MCDB105-M199
    mcdb105_m199 = make_medium(mcdb105, m199)
    mcdb105_m199.to_excel(writer, sheet_name='MCDB105-M199', index=False)

    # RPMI-EMEM
    rpmi_emem = make_medium(rpmi, emem)
    rpmi_emem.to_excel(writer, sheet_name='RPMI-EMEM', index=False)

    # RPMI-F12
    rpmi_f12 = make_medium(rpmi, f12)
    rpmi_f12.to_excel(writer, sheet_name='RPMI-F12', index=False)

    writer.save()

#make_medium_conditions()

def map_recon1_xchange_to_medium():
    """
    """
    book = load_workbook('./Medium_conditions.xlsx')
    writer = pd.ExcelWriter('./Medium_conditions.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    metabolite_map = pd.read_csv('./medium_component_mapped_to_recon1.csv')

    for medium in writer.sheets:
        if medium == "Summary":
            pass
        else:
            df = pd.read_excel('./Medium_conditions.xlsx', sheet_name=medium)
            merged_df = pd.merge(
                df, metabolite_map, how='inner', left_on=['Components'], right_on=['Original IDs'])
            merged_df = merged_df.drop_duplicates(keep='first')
            merged_df.to_excel(
                writer, sheet_name=medium, index=False, header=True)
    writer.save()

#map_recon1_xchange_to_medium()

def merge_duplicate_metabolites():
    book = load_workbook('./Medium_conditions.xlsx')
    writer = pd.ExcelWriter('./Medium_conditions.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    for medium in writer.sheets:
        if medium == "Summary":
            pass
        else:
            df = pd.read_excel('./Medium_conditions.xlsx', sheet_name=medium)
            df = df.groupby('Metabolite Name').mean().reset_index()
            df.to_csv('tmpx.csv')

merge_duplicate_metabolites()

def calculate_alpha():
    """
    """
    book = load_workbook(r'./Medium_conditions.xlsx')
    writer = pd.ExcelWriter(r'./Medium_conditions.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    rpmi = pd.read_excel(r'./Medium_conditions.xlsx',
                         sheet_name='RPMI', index_col='Components')

    for medium in writer.sheets:
        if medium == "Summary":
            pass
        else:
            other_medium = pd.read_excel(
                r'./Medium_conditions.xlsx', sheet_name=medium, index_col='Components')
            other_medium['alpha'] = other_medium['g/L'].divide(
                rpmi['g/L'], axis='index', fill_value=0)
            other_medium['alpha'] = other_medium['alpha'].replace(np.inf, 10)
            other_medium.to_excel(writer, sheet_name=medium, index=True)
    writer.save()


#calculate_alpha()





def scale_LB():
    """
    """
    book = load_workbook(r'./Medium_conditions.xlsx')
    writer = pd.ExcelWriter(r'./Medium_conditions.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    default_uptake_rate = pd.read_excel(
        r'./Medium_conditions.xlsx', sheet_name='RPMI', index_col='Components')

    for medium in writer.sheets:
        if medium == "Summary":
            pass
        else:
            uptake_rate_to_change = pd.read_excel(
                r'./Medium_conditions.xlsx', sheet_name=medium, index_col='Components')

            uptake_rate_to_change['Adjusted LB'] = default_uptake['LB'].multiply(
                uptake_rate_to_change['alpha'], axis=0, fill_value=1.00)
            uptake_rate_to_change.to_excel(
                writer, sheet_name=medium, index=True)
    writer.save()


#scale_LB()
